import openpyxl
import os
os.chdir('C:\\Users\\monis\\Desktop\\Pyprojects\\Learning\\Automate_the_Boring_Stuff')
print('path changed')
wb=openpyxl.load_workbook('example.xlsx')
type(wb)
for sheet in wb:           # go over all sheets
    print(sheet.title)
sheet=["Sheet3"]
type(sheet)
sheet = wb.active
print(sheet.title)
x=sheet.get_column_letter(sheet.max_column)
print(x)
